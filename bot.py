import os
import random
from datetime import datetime

from openpyxl import Workbook, load_workbook

from telegram import (
    Update,
    ReplyKeyboardMarkup,
    InputFile
)

from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

# =====================================================
# НАСТРОЙКИ
# =====================================================

TOKEN = "8429654652:AAHEDXDt8nT_mczfubd4rxf-OkUTFoP0wFk"

ADMIN_ID = 8348404226

# =====================================================
# ПРЕДМЕТЫ И КЛАССЫ
# =====================================================

subjects = {
    "Математика": ["5", "6"],
    "Естествознание": ["5", "6"],

    "Алгебра": ["7", "8", "9", "10", "11"],
    "Геометрия": ["7", "8", "9", "10", "11"],
    "Физика": ["7", "8", "9", "10", "11"],
    "Химия": ["7", "8", "9", "10", "11"],
    "Биология": ["7", "8", "9", "10", "11"],

    "Информатика": ["5", "6", "7", "8", "9", "10", "11"],
    "География": ["5", "6", "7", "8", "9", "10", "11"],
    "Русский язык": ["5", "6", "7", "8", "9", "10", "11"],
    "Русская литература": ["5", "6", "7", "8", "9", "10", "11"],
    "Казахский язык": ["5", "6", "7", "8", "9", "10", "11"],
    "Казахская литература": ["5", "6", "7", "8", "9", "10", "11"],
    "История Казахстана": ["5", "6", "7", "8", "9", "10", "11"],
    "Всемирная история": ["5", "6", "7", "8", "9", "10", "11"],
}

# =====================================================
# ГЕНЕРАЦИЯ ВОПРОСОВ
# =====================================================

question_bank = {}

def generate_questions(subject, grade):
    questions = []

    for i in range(1, 51):
        questions.append({
            "question": f"[{subject} {grade} класс] Вопрос №{i}",
            "options": [
                f"Ответ A{i}",
                f"Ответ B{i}",
                f"Ответ C{i}"
            ],
            "answer": f"Ответ A{i}"
        })

    return questions

for subject, grades in subjects.items():

    question_bank[subject] = {}

    for grade in grades:
        question_bank[subject][grade] = generate_questions(
            subject,
            grade
        )

# =====================================================
# ДАННЫЕ ПОЛЬЗОВАТЕЛЕЙ
# =====================================================

user_data = {}

# =====================================================
# EXCEL
# =====================================================

def save_to_excel(data, score, subject, grade):

    file_name = "results.xlsx"

    if not os.path.exists(file_name):

        wb = Workbook()
        ws = wb.active

        headers = [
            "ФИО",
            "Предмет",
            "Класс",
            "Дата и время"
        ] + [f"В{i}" for i in range(1, 16)] + ["Балл"]

        ws.append(headers)

        wb.save(file_name)

    wb = load_workbook(file_name)
    ws = wb.active

    current_time = datetime.now().strftime(
        "%d.%m.%Y %H:%M"
    )

    row = [
        data[0][0],
        subject,
        grade,
        current_time
    ]

    for ans in data:
        row.append(
            "✔" if ans[4] else "❌"
        )

    row.append(score)

    ws.append(row)

    wb.save(file_name)

# =====================================================
# START
# =====================================================

async def start(update: Update,
                context: ContextTypes.DEFAULT_TYPE):

    user_id = update.message.chat_id

    user_data[user_id] = {

        "name": None,

        "subject": None,
        "grade": None,

        "score": 0,
        "q": 0,

        "answers": [],
        "questions": [],

        "results_mode": False,
        "results_subject": None
    }

    keyboard = [[s] for s in subjects.keys()]

    await update.message.reply_text(
        "Добро пожаловать в систему тестирования!\n\nВыберите предмет:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

# =====================================================
# RESULTS
# =====================================================

async def send_results(update: Update,
                       context: ContextTypes.DEFAULT_TYPE):

    if update.message.chat_id != ADMIN_ID:

        await update.message.reply_text(
            "Нет доступа"
        )

        return

    user_id = update.message.chat_id

    user_data[user_id]["results_mode"] = True

    keyboard = [[s] for s in subjects.keys()]

    await update.message.reply_text(
        "Выберите предмет:",
        reply_markup=ReplyKeyboardMarkup(
            keyboard,
            resize_keyboard=True
        )
    )

# =====================================================
# HANDLE MESSAGE
# =====================================================

async def handle_message(update: Update,
                         context: ContextTypes.DEFAULT_TYPE):

    user_id = update.message.chat_id
    text = update.message.text

    if user_id not in user_data:
        await start(update, context)
        return

    data = user_data[user_id]

    # =================================================
    # РЕЖИМ RESULTS
    # =================================================

    if data.get("results_mode"):

        if data["results_subject"] is None:

            if text in subjects:

                data["results_subject"] = text

                keyboard = [
                    [g]
                    for g in subjects[text]
                ]

                await update.message.reply_text(
                    "Выберите класс:",
                    reply_markup=ReplyKeyboardMarkup(
                        keyboard,
                        resize_keyboard=True
                    )
                )

                return

        else:

            selected_subject = data["results_subject"]
            selected_grade = text

            source_file = "results.xlsx"

            if not os.path.exists(source_file):

                await update.message.reply_text(
                    "Файл результатов не найден"
                )

                return

            wb = load_workbook(source_file)
            ws = wb.active

            new_wb = Workbook()
            new_ws = new_wb.active

            headers = [
                cell.value
                for cell in ws[1]
            ]

            new_ws.append(headers)

            for row in ws.iter_rows(
                min_row=2,
                values_only=True
            ):

                subject = str(row[1])
                grade = str(row[2])

                if (
                    subject == selected_subject
                    and grade == selected_grade
                ):

                    new_ws.append(row)

            filtered_file = (
                f"{selected_subject}_{selected_grade}.xlsx"
            )

            new_wb.save(filtered_file)

            with open(filtered_file, "rb") as file:

                await update.message.reply_document(
                    document=InputFile(file)
                )

            data["results_mode"] = False
            data["results_subject"] = None

            return

    # =================================================
    # ВЫБОР ПРЕДМЕТА
    # =================================================

    if data["subject"] is None:

        if text in subjects:

            data["subject"] = text

            grades = subjects[text]

            keyboard = [[g] for g in grades]

            await update.message.reply_text(
                "Выберите класс:",
                reply_markup=ReplyKeyboardMarkup(
                    keyboard,
                    resize_keyboard=True
                )
            )

            return

    # =================================================
    # ВЫБОР КЛАССА
    # =================================================

    if data["grade"] is None:

        if text in subjects[data["subject"]]:

            data["grade"] = text

            await update.message.reply_text(
                "Введите ваше ФИО:"
            )

            return

    # =================================================
    # ВВОД ФИО
    # =================================================

    if data["name"] is None:

        data["name"] = text

        all_questions = question_bank[
            data["subject"]
        ][
            data["grade"]
        ]

        data["questions"] = random.sample(
            all_questions,
            15
        )

        await send_question(update, context)

        return

    # =================================================
    # ОТВЕТЫ
    # =================================================

    q_index = data["q"]

    if q_index < len(data["questions"]):

        q = data["questions"][q_index]

        correct = q["answer"]

        is_correct = text == correct

        if is_correct:
            data["score"] += 1

        data["answers"].append([
            data["name"],
            q["question"],
            text,
            correct,
            is_correct
        ])

        data["q"] += 1

        await send_question(update, context)

# =====================================================
# ОТПРАВКА ВОПРОСОВ
# =====================================================

async def send_question(update: Update,
                        context: ContextTypes.DEFAULT_TYPE):

    user_id = update.message.chat_id

    data = user_data[user_id]

    q_index = data["q"]

    if q_index < len(data["questions"]):

        q = data["questions"][q_index]

        keyboard = [
            [opt]
            for opt in q["options"]
        ]

        await update.message.reply_text(
            q["question"],
            reply_markup=ReplyKeyboardMarkup(
                keyboard,
                resize_keyboard=True
            )
        )

    else:

        await finish_test(update, context)

# =====================================================
# ЗАВЕРШЕНИЕ
# =====================================================

async def finish_test(update: Update,
                      context: ContextTypes.DEFAULT_TYPE):

    user_id = update.message.chat_id

    data = user_data[user_id]

    save_to_excel(
        data["answers"],
        data["score"],
        data["subject"],
        data["grade"]
    )

    percent = round(
        data["score"] / 15 * 100
    )

    await update.message.reply_text(
        f"Тест завершён!\n\n"
        f"Предмет: {data['subject']}\n"
        f"Класс: {data['grade']}\n"
        f"Балл: {data['score']}/15\n"
        f"Процент: {percent}%"
    )

# =====================================================
# RUN
# =====================================================

app = ApplicationBuilder().token(TOKEN).build()

app.add_handler(
    CommandHandler("start", start)
)

app.add_handler(
    CommandHandler("results", send_results)
)

app.add_handler(
    MessageHandler(
        filters.TEXT & ~filters.COMMAND,
        handle_message
    )
)

print("Бот запущен")

app.run_polling()
